"""Spreadsheet intake parser and engine-answer builder."""

from __future__ import annotations

from datetime import datetime
from pathlib import PurePosixPath
from xml.etree import ElementTree as ET
import zipfile
from typing import Dict, List, Tuple

from heuristics import infer_q27_baseline, infer_q28_safety
from keyword_extractor import extract_keywords_with_variability

TIMESTAMP_HEADER = "Timestamp"
FULL_NAME_HEADER = "Full_Name"
EMAIL_HEADER = "2. Email address"
LANGUAGE_HEADER = "Language"

Q13_HEADER = "Section 4. Market & Role Outlook\n\n13. How do you see the future demand for the target role/field over the next 3–5 years?"
Q15_HEADER = "15. How differentiated do you think your profile would be in this role/field?"
Q16_HEADER = "Section 5. Company / Team Stability & Structure\n\n16. How confident are you about the company’s financial health and long-term viability for this role?"
Q17_HEADER = "17. Overall, how stable and predictable does this company / team feel in terms of strategy, leadership, and workflow?"
Q18_HEADER = "18. Has the company (or target business unit) gone through major restructuring or leadership changes in the last 2–3 years?"
Q19_HEADER = "Section 6. Informal / Workflow / Policy Signals\n\n19. Do you already have a manager or senior colleague in the target role/team who would actively support your success?"
Q20_HEADER = "20. How clearly do you understand the decision-making and reporting lines in the target role?"
Q21_HEADER = "21. From what you can see, how would you describe the day-to-day workflow and pace in the target role?"
Q22_HEADER = "22. How exposed does this role feel to policy / regulatory / market shocks in the next 3–5 years?"
Q23_HEADER = "Section 7. Personal Fit & Temperament\n\n23. In big career decisions, what do you naturally optimize for first?"
Q24_HEADER = "24. How comfortable are you with taking visible risks in your career (moves that could clearly fail)?"
Q25_HEADER = "25. When you commit to a new path, how do you usually behave?"
Q27_HEADER = "Section 8. Current Reality & Safety Nets\n\n27. If you do nothing and stay where you are for the next 12–18 months, what is the most realistic scenario \n(Paragraph, 3–5 sentences)"
Q28_HEADER = "28. What realistic Plan B or safety nets do you have if this move does not work out? \n(Paragraph)"

REQUIRED_ENGINE_HEADERS = [
    Q13_HEADER,
    Q15_HEADER,
    Q16_HEADER,
    Q17_HEADER,
    Q18_HEADER,
    Q19_HEADER,
    Q20_HEADER,
    Q21_HEADER,
    Q22_HEADER,
    Q23_HEADER,
    Q24_HEADER,
    Q25_HEADER,
    Q27_HEADER,
    Q28_HEADER,
]

REQUIRED_ENGINE_TOKENS = [
    "13.",
    "15.",
    "16.",
    "17.",
    "18.",
    "19.",
    "20.",
    "21.",
    "22.",
    "23.",
    "24.",
    "25.",
    "27.",
    "28.",
]


def _normalize_string(value: object) -> str:
    return str(value or "").strip()


def _contains(text: str, *needles: str) -> bool:
    hay = text.casefold()
    return any(needle.casefold() in hay for needle in needles)


def _parse_timestamp(value: object):
    if isinstance(value, datetime):
        return value
    text = _normalize_string(value)
    if not text:
        return None

    candidates = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d",
    ]
    for fmt in candidates:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _map_q13(value: str) -> str:
    if _contains(value, "declining"):
        return "declining"
    if _contains(value, "mixed", "unclear"):
        return "mixed"
    if _contains(value, "growing", "increasing"):
        return "growing"
    return "mixed"


def _map_q15(value: str) -> str:
    if _contains(value, "many people", "similar profiles"):
        return "many_similar"
    if _contains(value, "some differentiation", "competitive"):
        return "some_diff"
    if _contains(value, "highly unique", "very unique"):
        return "highly_unique"
    return "some_diff"


def _map_q16(value: str) -> str:
    if _contains(value, "low"):
        return "low"
    if _contains(value, "medium"):
        return "medium"
    if _contains(value, "high"):
        return "high"
    return "medium"


def _map_q17(value: str) -> str:
    if _contains(value, "unstable"):
        return "unstable"
    if _contains(value, "somewhat stable", "noticeable changes", "uncertainty", "mixed"):
        return "mixed"
    if _contains(value, "stable and predictable"):
        return "stable"
    return "mixed"


def _map_q18(value: str) -> str:
    if _contains(value, "no major"):
        return "no_changes"
    if _contains(value, "multiple", "significant", "major") or _contains(value, "yes"):
        return "major_changes"
    if _contains(value, "some"):
        return "some_changes"
    if _contains(value, "no"):
        return "no_changes"
    return "some_changes"


def _map_q19(value: str) -> str:
    if _contains(value, "no", "none"):
        return "none"
    if _contains(value, "maybe", "not sure", "some goodwill"):
        return "weak"
    if _contains(value, "strong sponsor", "yes"):
        return "strong"
    return "weak"


def _map_q20(value: str) -> str:
    if _contains(value, "not clear"):
        return "not_clear"
    if _contains(value, "somewhat clear"):
        return "somewhat_clear"
    if _contains(value, "very clear", "clear"):
        return "clear"
    return "somewhat_clear"


def _map_q21(value: str) -> str:
    if _contains(value, "very slow", "chaotic"):
        return "very_slow_or_chaotic"
    if _contains(value, "mixed"):
        return "mixed"
    if _contains(value, "clear", "sustainable"):
        return "clear_and_sustainable"
    return "mixed"


def _map_q22(value: str) -> str:
    if _contains(value, "highly exposed"):
        return "high_exposure"
    if _contains(value, "manageable", "some exposure"):
        return "medium_exposure"
    if _contains(value, "low exposure", "protected"):
        return "low_exposure"
    return "medium_exposure"


def _map_q23(value: str) -> str:
    if _contains(value, "maximize", "upside"):
        return "maximize_upside"
    if _contains(value, "balance"):
        return "balance"
    if _contains(value, "protecting the downside", "downside"):
        return "protect_downside"
    return "balance"


def _map_q24(value: str) -> str:
    if _contains(value, "not comfortable"):
        return "not_comfortable"
    if _contains(value, "somewhat"):
        return "somewhat_comfortable"
    if _contains(value, "comfortable"):
        return "comfortable"
    return "somewhat_comfortable"


def _map_q25(value: str) -> str:
    if _contains(value, "impulsive"):
        return "impulsive"
    if _contains(value, "gradually", "small, reversible", "testing"):
        return "gradual"
    if _contains(value, "deliberate", "carefully"):
        return "deliberate"
    return "gradual"


def _non_empty_row(values: Tuple[object, ...]) -> bool:
    return any(_normalize_string(cell) for cell in values)


def _row_to_dict(headers: List[str], values: Tuple[object, ...]) -> Dict[str, object]:
    row = {}
    for idx, header in enumerate(headers):
        row[header] = values[idx] if idx < len(values) else ""
    return row


def _xlsx_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha()).upper()
    result = 0
    for ch in letters:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return max(0, result - 1)


def _load_shared_strings(zf: zipfile.ZipFile) -> List[str]:
    try:
        raw = zf.read("xl/sharedStrings.xml")
    except KeyError:
        return []

    root = ET.fromstring(raw)
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    strings = []
    for si in root.findall("a:si", ns):
        text_parts = [node.text or "" for node in si.findall(".//a:t", ns)]
        strings.append("".join(text_parts))
    return strings


def _resolve_sheet_path(zf: zipfile.ZipFile, sheet_name: str) -> str:
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    ns_wb = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    ns_rel = {"a": "http://schemas.openxmlformats.org/package/2006/relationships"}

    rid_to_target = {}
    for rel in rels_root.findall("a:Relationship", ns_rel):
        rid = rel.attrib.get("Id", "")
        target = rel.attrib.get("Target", "")
        if rid and target:
            rid_to_target[rid] = target

    for sheet in wb_root.findall("a:sheets/a:sheet", ns_wb):
        name = sheet.attrib.get("name", "")
        rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
        if name == sheet_name and rid in rid_to_target:
            target = rid_to_target[rid]
            cleaned = target.lstrip("/")
            full = PurePosixPath("xl") / PurePosixPath(cleaned)
            return str(full)

    raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")


def _read_sheet_rows_via_xml(xlsx_path: str, sheet_name: str) -> List[Tuple[object, ...]]:
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        shared_strings = _load_shared_strings(zf)
        sheet_path = _resolve_sheet_path(zf, sheet_name)
        sheet_root = ET.fromstring(zf.read(sheet_path))

    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rows: List[Tuple[object, ...]] = []

    for row in sheet_root.findall("a:sheetData/a:row", ns):
        row_values: Dict[int, object] = {}
        max_idx = -1
        for cell in row.findall("a:c", ns):
            ref = cell.attrib.get("r", "")
            if not ref:
                continue
            idx = _xlsx_col_index(ref)
            max_idx = max(max_idx, idx)
            ctype = cell.attrib.get("t", "")
            value_node = cell.find("a:v", ns)
            inline_node = cell.find("a:is/a:t", ns)
            text = value_node.text if value_node is not None else None

            value: object = ""
            if ctype == "s" and text is not None:
                try:
                    value = shared_strings[int(text)]
                except Exception:
                    value = ""
            elif ctype == "inlineStr":
                value = inline_node.text if inline_node is not None else ""
            elif ctype == "b" and text is not None:
                value = "TRUE" if text == "1" else "FALSE"
            elif text is not None:
                value = text

            row_values[idx] = value

        if max_idx >= 0:
            expanded = tuple(row_values.get(i, "") for i in range(max_idx + 1))
            rows.append(expanded)

    return rows


def read_latest_row(xlsx_path: str, sheet: str = "Scores") -> Dict[str, object]:
    rows: List[Tuple[object, ...]]
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(xlsx_path, data_only=True)
        if sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in workbook.")
        ws = workbook[sheet]
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        rows = _read_sheet_rows_via_xml(xlsx_path, sheet)
    if not rows:
        raise ValueError("Sheet is empty.")

    raw_headers = ["" if h is None else str(h) for h in rows[0]]
    if not any(raw_headers):
        raise ValueError("Header row is empty.")

    normalized_headers = [h.strip() for h in raw_headers]
    data_rows = [r for r in rows[1:] if _non_empty_row(r)]
    if not data_rows:
        raise ValueError("No non-empty response rows found.")

    timestamp_idx = None
    if TIMESTAMP_HEADER in raw_headers:
        timestamp_idx = raw_headers.index(TIMESTAMP_HEADER)
    elif TIMESTAMP_HEADER in normalized_headers:
        timestamp_idx = normalized_headers.index(TIMESTAMP_HEADER)

    selected_row = data_rows[-1]
    if timestamp_idx is not None:
        parsed: List[Tuple[datetime, Tuple[object, ...]]] = []
        for row in data_rows:
            if timestamp_idx < len(row):
                dt = _parse_timestamp(row[timestamp_idx])
                if dt is not None:
                    parsed.append((dt, row))
        if parsed:
            parsed.sort(key=lambda item: item[0])
            selected_row = parsed[-1][1]

    return _row_to_dict(raw_headers, selected_row)


def _require_headers(row: Dict[str, object], headers: List[str]) -> None:
    missing = [header for header in headers if header not in row]
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))


def _find_header_by_token(row: Dict[str, object], token: str) -> str:
    normalized_token = token.strip().casefold()
    for header in row.keys():
        if normalized_token in str(header).casefold():
            return str(header)
    return ""


def _resolve_required_headers(row: Dict[str, object]) -> Dict[str, str]:
    resolved: Dict[str, str] = {}
    missing: List[str] = []

    for token in REQUIRED_ENGINE_TOKENS:
        found = _find_header_by_token(row, token)
        if found:
            resolved[token] = found
        else:
            missing.append(token)

    if missing:
        raise ValueError(
            "Missing required column token(s): "
            + ", ".join(missing)
            + f" (available headers: {len(row.keys())})"
        )

    return resolved


def infer_language(row: Dict[str, object], lang_override: str = "") -> str:
    if lang_override:
        return "ko" if lang_override.strip().casefold() in {"ko", "kr", "korean", "한국어", "ko-kr"} else "en"

    raw = _normalize_string(row.get(LANGUAGE_HEADER, ""))
    if raw and raw.casefold() in {"ko", "kr", "korean", "한국어", "ko-kr"}:
        return "ko"
    return "en"


def build_engine_answers(row: Dict[str, object], lang: str = "en") -> Dict[str, str]:
    resolved = _resolve_required_headers(row)

    q27_text = _normalize_string(row.get(resolved["27."], ""))
    q28_text = _normalize_string(row.get(resolved["28."], ""))

    return {
        "q13_outlook": _map_q13(_normalize_string(row.get(resolved["13."], ""))),
        "q15_diff": _map_q15(_normalize_string(row.get(resolved["15."], ""))),
        "q16_health": _map_q16(_normalize_string(row.get(resolved["16."], ""))),
        "q17_stability": _map_q17(_normalize_string(row.get(resolved["17."], ""))),
        "q18_change": _map_q18(_normalize_string(row.get(resolved["18."], ""))),
        "q19_sponsor": _map_q19(_normalize_string(row.get(resolved["19."], ""))),
        "q20_clarity": _map_q20(_normalize_string(row.get(resolved["20."], ""))),
        "q21_workflow": _map_q21(_normalize_string(row.get(resolved["21."], ""))),
        "q22_exposure": _map_q22(_normalize_string(row.get(resolved["22."], ""))),
        "q23_temperament": _map_q23(_normalize_string(row.get(resolved["23."], ""))),
        "q24_risk_comfort": _map_q24(_normalize_string(row.get(resolved["24."], ""))),
        "q25_commit_style": _map_q25(_normalize_string(row.get(resolved["25."], ""))),
        "q27_baseline": infer_q27_baseline(q27_text, lang),
        "q28_safety": infer_q28_safety(q28_text, lang),
    }


def parse_pipeline_inputs(row: Dict[str, object], lang: str) -> Dict[str, object]:
    answers = build_engine_answers(row, lang)
    resolved = _resolve_required_headers(row)

    q27_text = _normalize_string(row.get(resolved["27."], ""))
    q28_text = _normalize_string(row.get(resolved["28."], ""))

    baseline = extract_keywords_with_variability(q27_text, lang)
    safety = extract_keywords_with_variability(q28_text, lang)
    combined = extract_keywords_with_variability(f"{q27_text}\n{q28_text}".strip(), lang)

    return {
        "answers": answers,
        "q27_raw": q27_text,
        "q28_raw": q28_text,
        "baseline_keywords": baseline,
        "safetynet_keywords": safety,
        "combined_keywords": combined,
    }
